from xml.etree.ElementTree import QName
from openpyxl import load_workbook
import os
import requests
import re


class convert_savepatch:
    def __init__(self):
        self.eu_db = ""
        self.us_db = ""
        # interpreter working dir = apollo-vita
        # if interpreter working dir = region converter folder, then script_path should be empty = ""
        self.script_path = "region convertor" 
        self.db_path = f"{self.script_path}\\db"
        self.save_path = f"{self.script_path}\\converted"
        self.savepatch = []

    def fetch_savepatch_info(self, savepatch: str):
        ####################################################
        ##  Fetch region and game title from .savepatch
        ####################################################
        info = {"title": "", "region": ""}

        with open(f"{self.script_path}\{savepatch}") as file:
            first_line = file.readlines()[0].strip()
            index_of_region = first_line.find("(")
            region = first_line[index_of_region + 1 : -1]
            if "EU" in region:
                info["region"] = "EU"
            elif "US" in region:
                info["region"] = "US"
            else:
                return None

            title = first_line[first_line.find(" ") : index_of_region]
            info["title"] = title

        return info
                    
    def prepare_db(self):
        self.us_db = load_workbook(f"{self.db_path}\\US.xlsx").active
        self.eu_db = load_workbook(f"{self.db_path}\\EU.xlsx").active

    def generate_savepatch(self, id_in_dif_region, file_name, region):
        ########################################################################
        ##  change region title on first line and generate a new .savepatch
        #######################################################################
        id_in_dif_region += ".savepatch"
        if region == "EU":
            region = "USA"
        else:
            region = "EUR"
        with open(self.script_path + "\\" + file_name) as read:
            with open(self.script_path + "\\converted\\" + id_in_dif_region, "w+") as write:
                data = read.readlines()
                line = data[0]
                line = f"{line[ : line.find('(')]}({region})\n"
                data[0] = line
                
                write.write("".join(data))
                print(f"Generated {file_name}.")
                
        return True
                
    def converter(self, region, id, title):
        #############################################################
        ##  if title in EU, find its equavalent in US and vice versa
        #############################################################
        eu_first_column = self.eu_db["A"]
        eu_second_column = self.eu_db["B"]
        us_first_column = self.us_db["A"]
        us_second_column = self.us_db["B"]
        id_in_dif_region = None

        if region == "EU":
            get_id_from = us_first_column
            get_title_from = us_second_column
            
        elif region == "US":
            get_id_from = eu_first_column
            get_title_from = eu_second_column
            
        for index, fetch_title in enumerate(get_title_from):
            ####################################################
            ##  Find the opposite db for the same title
            ####################################################
            if fetch_title.value == None:
                # end of excel
                break
            
            fetch_title = fetch_title.value.lower().strip()
            title = title.lower().strip()

            ##################################################################
            ##  Check only titles starting the same char, then the whole title
            ##################################################################
            id_in_dif_region = None
            if fetch_title[0] == title[0]:
                if title in fetch_title:
                    id_in_dif_region = get_id_from[index].value

                    print(f"Found: ({title}) in offline database.\n")

                    self.generate_savepatch(id_in_dif_region, id, region)
                    break

        if id_in_dif_region == None:
            ########################################################################
            ##      Online search using serialstation.com
            #######################################################################
            print(f"Not found: ({title}) in offline database")
            print(f"Searching from online database, wish me luck!...")

            search = ""
            for i in title:
                if i == " ":
                    search += "+"
                else:
                    if "&" in i:
                        search += "and"
                    else:
                        search += i
            data = requests.get("https://www.serialstation.com/search/?search=" + search).text.split('\n')
            components = [x for x in data if '<span class="badge bg-secondary">' in x and '>Original<' not in x and '>None<' not in x]
            ids = []
            us = "North America"
            eu = ("United Kingdom", "Italy", "Germany", "Australia", "Spain", "Europe")
            found = None
            id_found = ""

            for index, c in enumerate(components):
                if '<span class="badge bg-secondary">' in c and 'PS3' not in c:
                    c = c.replace('<span class="badge bg-secondary">', '')
                    c = c.replace('</span>', '')
                    ids.append(c)

            for index, i in enumerate(ids):
                if region == "EU":
                    if us in i:
                        id_found = ids[index+1].replace("-", "")
                        found = self.generate_savepatch(id_found, id, region)
                        break

                elif region == "US":
                    for available_region in eu:
                        if i == available_region:
                            id_found = ids[index+1].replace("-", "")
                            found = self.generate_savepatch(id_found, id, region)
                            break
                        break

            if found:
                print(f"Found: ({title}) = ({id_found}) from the online database.\n")
                
            else:
                print(f"No luck, I couldn't find it from online database :(")
                if region == "EU" :id_in_dif_region = "USA" 
                elif region == "US" :id_in_dif_region = "EUR" 
                if len(ids) > 0:
                    for i in ids:
                        if len(i) == 3: # Game version i.e. PSV, PS3
                            print(f"* {i}:")
                            print("\n")
                        elif len(i) == 10: # Game id
                            print(i.replace("-", ""))
                            print("_____________________________________")
                        else:
                            print(i)
                    print(f"\nFound these game ids. Pick one that mention 'PSV' if found, if not available above consider googling")
                else:
                    print(f"Cannot find ({title}) at serialstation.com")
                id_from_user = input(f"Paste ({id_in_dif_region}) id for ({title}) & I'll generate the file or type 's' to (skip): ")
                if id_from_user != "s":
                    self.generate_savepatch(id_from_user, id, region)
                
    def start(self):
        self.prepare_db()

        dir = os.listdir(self.script_path)
        self.savepatch = [file_name for file_name in dir if ".savepatch" in file_name]
        
        for game_id in self.savepatch:
            game = self.fetch_savepatch_info(game_id)
            if game:
                game_region, game_title = game["region"], game["title"]
                self.converter(game_region, game_id, game_title)

convert_savepatch().start()
